from flask import Flask, render_template, request, redirect, url_for, flash, session, send_from_directory, send_file
from openpyxl import load_workbook, Workbook
from io import BytesIO
import os
from dotenv import load_dotenv
import secrets
import socket
from datetime import datetime

app = Flask(__name__)
app.secret_key = secrets.token_hex(16)  # Generate a secure random key

# Initialize session tracking
def init_session():
    if 'user_id' not in session:
        # Use PC name as user ID and store it persistently
        session.permanent = True  # Make session permanent
        session['user_id'] = socket.gethostname()
        session['submitted_subjects'] = []

def validate_submission(subject):
    """Check if feedback already exists in the Excel file"""
    try:
        wb = load_workbook('feedback.xlsx')
        tracking_sheet = wb['_submissions_tracking']
        user_id = socket.gethostname()  # Always use current hostname
        
        # Check Excel file for existing submission
        for row in tracking_sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == user_id and row[1] == subject:
                return False
        return True
    except Exception:
        return True  # If file doesn't exist or other error, allow submission

def get_questions():
    return [
        "How would you rate the overall teaching quality?",
        "How clearly was the subject matter explained?",
        "How effective was the use of examples and illustrations?",
        "How was the pace of the class?",
        "How effectively were your doubts resolved?",
        "How would you rate the study material quality?",
        "How helpful were the practice problems?",
        "Was faculty support available when needed?",
        "How confident do you feel about the exam?"
    ]

# Load configuration
load_dotenv()
ADMIN_PASSWORD = os.getenv('ADMIN_PASSWORD', 'admin123')  # Default password if .env not present

# Helper functions
def load_subjects():
    try:
        with open('subjects.txt', 'r') as f:
            return [line.strip() for line in f if line.strip()]
    except FileNotFoundError:
        return []

def save_feedback(subject, responses, user_id):
    filename = 'feedback.xlsx'
    if os.path.exists(filename):
        wb = load_workbook(filename)
    else:
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

    # Create or get the submissions tracking sheet
    tracking_sheet_name = '_submissions_tracking'
    if tracking_sheet_name not in wb.sheetnames:
        tracking_sheet = wb.create_sheet(tracking_sheet_name)
        tracking_sheet.append(['user_id', 'subject', 'timestamp'])
    else:
        tracking_sheet = wb[tracking_sheet_name]

    # Add submission record
    tracking_sheet.append([user_id, subject, datetime.now().isoformat()])

    if subject not in wb.sheetnames:
        wb.create_sheet(subject)
    
    sheet = wb[subject]
    
    # Add headers if sheet is empty
    if sheet.max_row == 1:
        # Get questions for headers
        questions = get_questions()
        headers = ['Computer Name'] + questions + ['Most Helpful Aspects', 'Suggestions']
        sheet.append(headers)
    
    # Add responses with user_id
    sheet.append([user_id] + responses)
    wb.save(filename)

def has_submitted_feedback(subject):
    """Check both session and Excel file for existing submissions"""
    # First check Excel file (source of truth)
    if not validate_submission(subject):
        # Update session to match Excel data
        if subject not in session.get('submitted_subjects', []):
            if 'submitted_subjects' not in session:
                session['submitted_subjects'] = []
            session['submitted_subjects'].append(subject)
        return True
    
    # Then check session
    if subject in session.get('submitted_subjects', []):
        return True

    return False

def get_question_options():
    return [
        ['Excellent', 'Good', 'Average', 'Needs Improvement'],
        ['Strongly Agree', 'Agree', 'Neutral', 'Disagree', 'Strongly Disagree'],
        ['Always', 'Often', 'Rarely', 'Never'],
        ['Too Fast', 'Just Right', 'Too Slow'],
        ['Yes, very effectively', 'Somewhat', 'Not effectively', 'Not applicable'],
        ['Excellent', 'Good', 'Average', 'Poor'],
        ['Yes', 'No', 'To some extent'],
        ['Yes', 'No', "Didn't need support"],
        ['Very Confident', 'Somewhat Confident', 'Not Confident']
    ]

def get_option_index(question_num, option_text):
    options = get_question_options()[question_num - 1]
    try:
        return options.index(option_text) + 1
    except ValueError:
        return 0

def calculate_average_ratings(responses):
    averages = []
    for i in range(9):  # First 9 questions are MCQ
        total = sum(get_option_index(i + 1, row[i]) for row in responses)
        avg = total / len(responses) if responses else 0
        averages.append(round(avg, 2))
    return averages

def get_rating_distribution(responses):
    distribution = [0] * 5  # Maximum 5 options in any question
    for response in responses:
        for i in range(9):  # Only process first 9 MCQ questions
            option_index = get_option_index(i + 1, response[i])
            if 1 <= option_index <= 5:  # Ensure index is within valid range
                distribution[option_index - 1] += 1
    return distribution

def get_sheet_data(sheet):
    """Get raw data from sheet for admin view"""
    raw_data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if all(cell is not None for cell in row):
            raw_data.append(row)
    return raw_data

# Routes
@app.route('/favicon.ico')
def favicon():
    return '', 204

@app.route('/')
def index():
    init_session()
    return render_template('index.html')

@app.route('/guidelines')
def guidelines():
    session['current_slide'] = session.get('current_slide', 1)
    return render_template('guidelines.html', slide=session['current_slide'])

@app.route('/next_slide')
def next_slide():
    session['current_slide'] = session.get('current_slide', 1) + 1
    if session['current_slide'] > 4:  # Total 4 slides
        session.pop('current_slide', None)  # Clear slide counter
        return redirect(url_for('select_subject'))
    return redirect(url_for('guidelines'))

@app.route('/prev_slide')
def prev_slide():
    session['current_slide'] = max(1, session.get('current_slide', 1) - 1)
    return redirect(url_for('guidelines'))

@app.route('/select_subject')
def select_subject():
    subjects = load_subjects()
    if not subjects:
        flash('No subjects available. Please contact administrator.')
        return redirect(url_for('index'))
    
    submitted_subjects = session.get('submitted_subjects', [])
    # Check if all subjects have been submitted
    if submitted_subjects and len(submitted_subjects) == len(subjects) and all(s in submitted_subjects for s in subjects):
        return redirect(url_for('completion_success'))
        
    return render_template('select_subject.html', subjects=subjects)

@app.route('/feedback/<subject>')
def feedback_form(subject):
    init_session()
    current_pc = socket.gethostname()
    session_pc = session.get('user_id')

    # Validate PC hasn't changed
    if session_pc and session_pc != current_pc:
        session.clear()
        flash('Session invalid - PC identifier changed.')
        return redirect(url_for('index'))

    if subject not in load_subjects():
        flash('Invalid subject selected.')
        return redirect(url_for('select_subject'))
    
    if not validate_submission(subject):
        flash('You have already submitted feedback for this subject from this computer.')
        return redirect(url_for('select_subject'))
        
    return render_template('feedback_form.html', subject=subject)

@app.route('/submit_feedback', methods=['POST'])
def submit_feedback():
    # Validate session and PC
    current_pc = socket.gethostname()
    session_pc = session.get('user_id')
    
    if not session_pc or session_pc != current_pc:
        session.clear()
        flash('Session invalid - PC identifier changed.')
        return redirect(url_for('index'))

    subject = request.form.get('subject')
    if not subject or subject not in load_subjects():
        flash('Invalid subject.')
        return redirect(url_for('select_subject'))

    responses = [request.form.get(f'q{i}') for i in range(1, 12)]
    
    if None in responses or '' in responses:
        flash('Please answer all questions.')
        return redirect(url_for('feedback_form', subject=subject))
    
    try:
        # Double-check Excel file for existing submission
        if not validate_submission(subject):
            flash('You have already submitted feedback for this subject from this computer.')
            return redirect(url_for('select_subject'))

        save_feedback(subject, responses, current_pc)
        session['submitted_subjects'] = session.get('submitted_subjects', []) + [subject]
        return redirect(url_for('thank_you'))
    except Exception as e:
        flash('Error saving feedback. Please try again.')
        return redirect(url_for('feedback_form', subject=subject))

@app.route('/thank_you')
def thank_you():
    return render_template('thank_you.html')

@app.route('/admin', methods=['GET', 'POST'])
def admin():
    # Clear any existing admin session on GET request
    if request.method == 'GET':
        session.pop('admin', None)
        return render_template('admin.html')
    
    # Handle POST request
    if request.form.get('password') == ADMIN_PASSWORD:
        session['admin'] = True
        session['admin_login_time'] = datetime.now().isoformat()
        return redirect(url_for('admin_dashboard'))
    
    flash('Invalid password.')
    return redirect(url_for('admin'))

@app.route('/admin/dashboard')
def admin_dashboard():
    if not session.get('admin'):
        flash('Please log in first.')
        return redirect(url_for('admin'))
    
    try:
        wb = load_workbook('feedback.xlsx')
        data = {}
        questions = get_questions()

        for sheet_name in wb.sheetnames:
            if sheet_name == '_submissions_tracking':
                continue

            data[sheet_name] = {
                'questions': questions,
                'raw_data': get_sheet_data(wb[sheet_name])
            }

        return render_template('admin_dashboard.html', data=data)
    except FileNotFoundError:
        flash('No feedback data available yet.')
        return render_template('admin_dashboard.html', data={})
    except Exception as e:
        flash(f'Error loading feedback data: {str(e)}')
        return render_template('admin_dashboard.html', data={})

@app.route('/completion_success')
def completion_success():
    subjects = load_subjects()
    submitted_subjects = session.get('submitted_subjects', [])
    
    # Verify all subjects are completed
    if not submitted_subjects or len(submitted_subjects) != len(subjects) or not all(s in submitted_subjects for s in subjects):
        return redirect(url_for('select_subject'))
    
    return render_template('completion_success.html', 
                         total_subjects=len(subjects),
                         completed_subjects=subjects)

@app.route('/logout')
def logout():
    session.clear()
    flash('Successfully logged out.')
    return redirect(url_for('index'))

@app.errorhandler(404)
def page_not_found(e):
    flash('Page not found.')
    return redirect(url_for('index'))

@app.route('/download_excel')
def download_excel():
    if not session.get('admin'):
        flash('Please log in as admin first.')
        return redirect(url_for('admin'))
    
    try:
        output = BytesIO()
        
        # Just send the existing Excel file with all sheets
        if os.path.exists('feedback.xlsx'):
            with open('feedback.xlsx', 'rb') as f:
                output.write(f.read())
            output.seek(0)
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'feedback_all_subjects_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            )
        else:
            flash('No feedback data available yet.')
            return redirect(url_for('admin_dashboard'))
    except Exception as e:
        flash(f'Error downloading file: {str(e)}')
        return redirect(url_for('admin_dashboard'))

@app.errorhandler(500)
def internal_error(e):
    flash('An internal error occurred. Please try again.')
    return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=True)
